#!/usr/bin/env python3
# =============================================================================
# TAJ ALAMEERA - Salla Product-Import XLSX Builder (openpyxl)
# -----------------------------------------------------------------------------
# Salla rejected our hand-zipped xlsx ("must be Microsoft Excel"). Its parser
# (PhpSpreadsheet) wants a genuinely valid OOXML file. This loads Salla's OWN
# downloaded template with openpyxl, writes our 28 product rows starting at
# row 3 (keeping Salla's header rows 1 & 2 + validation intact), and saves a
# clean xlsx Salla accepts.
#
# Run with the interpreter that HAS openpyxl:
#     /usr/bin/python3 build-import-xlsx.py
#
# Env:
#   TEMPLATE   Salla template path (default ~/Downloads/product-sample.xlsx)
#   OUT        output path (default ./salla-import.xlsx)
#   VAT_RATE   default 0 (owner adds VAT in Salla dashboard at checkout)
#   FORCE_STALE=1  allow stale proxy price
# =============================================================================
import os
import sys
import json
import urllib.request

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl missing. Run with /usr/bin/python3 (it has openpyxl 3.1.5).")

HERE = os.path.dirname(os.path.abspath(__file__))

PROXY_URL = "https://taj-gold-proxy.tajalamerahost.workers.dev/"
GRAM_MIN, GRAM_MAX = 150, 900
VAT_RATE = float(os.environ.get("VAT_RATE", "0"))
ALLOW_STALE = os.environ.get("FORCE_STALE") == "1"
TEMPLATE = os.environ.get("TEMPLATE", os.path.expanduser("~/Downloads/product-sample.xlsx"))
OUT = os.environ.get("OUT", os.path.join(HERE, "salla-import.xlsx"))
SEED = os.path.join(HERE, "seed-products.json")

# Salla rejects import when the category / product-type / brand columns carry
# values it wants set in the dashboard ("please delete category type brand").
# Default ON = leave C (تصنيف), F (نوع المنتج), T (الماركة) BLANK; owner assigns
# category + type in Salla after import (bulk-select). Set BLANK_META=0 to keep them.
BLANK_META = os.environ.get("BLANK_META", "1") != "0"

# Salla's official import doc: "Do not delete any columns from the template, even
# if you do not need them." So stripping columns is WRONG (default OFF). The doc's
# guidance for the "remove category/type/brand" prompt is to clear the VALUES
# (that's BLANK_META), keeping all columns intact. STRIP_META=1 kept only as an
# escape hatch; do not use it against the normal import flow.
STRIP_META = os.environ.get("STRIP_META", "0") == "1"

# profit tiers all 0 until owner sends real numbers (keep same as JS scripts)
PROFIT_TIERS = [(5, 0), (10, 0), (20, 0), (float("inf"), 0)]

# our seed category_ar -> Salla store's EXACT category name
CAT_MAP = {"أساور": "أساور", "سلاسل": "سلاسل وعقود", "خواتم": "خواتم"}


def round2(n):
    return round(n * 100) / 100


def round4(n):
    return round(n * 10000) / 10000


def profit_per_gram(w):
    for max_g, per in PROFIT_TIERS:
        if w <= max_g:
            return per
    return 0


def compute_price(gram, p):
    g = gram.get(p["karat"])
    if not isinstance(g, (int, float)):
        raise ValueError("no gram price karat %s" % p["karat"])
    gold = g * p["weight"]
    mk = p["making"]
    if mk["type"] == "per_gram":
        making = mk["value"] * p["weight"]
    elif mk["type"] == "fixed":
        making = mk["value"]
    else:
        raise ValueError("bad making type %s" % mk["type"])
    profit = profit_per_gram(p["weight"]) * p["weight"]
    return round2((gold + making + (p.get("stones") or 0) + profit) * (1 + VAT_RATE))


def get_gram():
    req = urllib.request.Request(PROXY_URL, headers={
        "Cache-Control": "no-store",
        "User-Agent": "Mozilla/5.0 (Macintosh; taj-import-builder)",
        "Accept": "application/json",
    })
    with urllib.request.urlopen(req, timeout=20) as r:
        d = json.loads(r.read().decode("utf-8"))
    if not d or not d.get("ok") or not d.get("prices"):
        raise ValueError("proxy not-ok / no prices")
    if d.get("stale") and not ALLOW_STALE:
        raise ValueError("proxy STALE. FORCE_STALE=1 to override.")
    prices = d["prices"]
    for k in ("24k", "21k", "18k"):
        v = float(prices[k])
        if v < GRAM_MIN or v > GRAM_MAX:
            raise ValueError("gram %s=%s outside bounds. ABORT." % (k, prices[k]))
    return {"24k": float(prices["24k"]), "21k": float(prices["21k"]),
            "18k": float(prices["18k"]), "time": d.get("time")}


KARAT_LABEL = {"24k": "24", "21k": "21", "18k": "18"}


def main():
    if not os.path.exists(TEMPLATE):
        sys.exit("template missing: %s (put Salla's product-sample.xlsx there or set TEMPLATE=)" % TEMPLATE)
    with open(SEED, "r", encoding="utf-8") as f:
        seed = json.load(f)

    gram = get_gram()
    print("Gram (SAR): 24k=%s 21k=%s 18k=%s time=%s" % (gram["24k"], gram["21k"], gram["18k"], gram["time"]))

    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.active

    # Salla rejects at upload with "delete category type brand". Physically REMOVE
    # those columns (تصنيف C=3, نوع المنتج F=6, الماركة T=20). Delete right-to-left
    # so earlier indices stay valid. After this the column letters shift, so the
    # write mapping below uses the POST-STRIP column map.
    if STRIP_META:
        for col in (20, 6, 3):          # T, F, C
            ws.delete_cols(col, 1)

    # Salla's template ships with MULTIPLE sample rows starting at row 3 (a product
    # row + variant rows) carrying gray "explanatory only" cells (FFd6d6d6) marked
    # ممنوع الكتابة plus sample junk (barcode, option colors, sizes, logo URL). The
    # official import doc says "Do not edit cells shaded in gray", and leftover sample
    # values in extra columns made Salla reject earlier. DELETE every sample row from
    # 3 to the end so our products write into clean rows with no residual junk in the
    # columns we don't set. Header rows 1 & 2 stay intact.
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    # Resolve columns by HEADER NAME (row 2) so the mapping survives the column
    # strip above (letters shift after delete_cols). Header text is trimmed.
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v is not None:
            hdr[str(v).strip()] = c

    def col(name):
        return hdr.get(name)

    C_TYPEROW = col("النوع")          # product/variant marker (A)
    C_NAME    = col("أسم المنتج")
    C_CAT     = col("تصنيف المنتج")   # None if stripped
    C_IMG     = col("صورة المنتج")
    C_ALT     = col("وصف صورة المنتج")
    C_PTYPE   = col("نوع المنتج")     # None if stripped
    C_PRICE   = col("سعر المنتج")
    C_DESC    = col("الوصف")
    C_SHIP    = col("هل يتطلب شحن؟")
    C_SKU     = col("رمز المنتج sku")
    C_WT      = col("الوزن")
    C_WTUNIT  = col("وحدة الوزن")

    def put(c, r, v):
        if c:
            ws.cell(row=r, column=c, value=v)

    row = 3  # rows 1 & 2 are Salla's headers
    ok = failed = 0
    for p in seed:
        try:
            price = compute_price(gram, p)
        except Exception as e:
            print("P%s %s: %s" % (p.get("folder"), p.get("name"), e))
            failed += 1
            continue
        karat = KARAT_LABEL.get(p["karat"], p["karat"])
        cat = CAT_MAP.get(p["category_ar"], p["category_ar"])
        if p.get("placeholder"):
            desc = "%s - عيار %s - وزن تقريبي %s جم. السعر محسوب من سعر الجرام المباشر. (بيانات مبدئية تُحدّث لاحقاً)" % (p["name"], karat, p["weight"])
        else:
            desc = "%s - عيار %s - %s جم." % (p["name"], karat, p["weight"])
        gallery = p.get("gallery") or []
        imgs = ",".join([u for u in (gallery if gallery else [p.get("hero")]) if u])
        kg = round4(p["weight"] / 1000.0)

        put(C_TYPEROW, row, "منتج")   # A النوع (product row marker; NOT product-type)
        put(C_NAME, row, p["name"])   # B أسم المنتج
        # C تصنيف المنتج: only when kept (STRIP_META=0 AND BLANK_META=0)
        if C_CAT and not BLANK_META:
            put(C_CAT, row, cat)
        put(C_IMG, row, imgs)         # D صورة المنتج
        put(C_ALT, row, p["name"])    # E وصف صورة المنتج (alt)
        # F نوع المنتج: only when kept
        if C_PTYPE and not BLANK_META:
            put(C_PTYPE, row, "منتج جاهز")
        put(C_PRICE, row, price)      # G سعر المنتج
        put(C_DESC, row, desc)        # H الوصف
        put(C_SHIP, row, "نعم")       # I هل يتطلب شحن؟
        put(C_SKU, row, p["sku"])     # J رمز المنتج sku
        put(C_WT, row, kg)            # R الوزن (kg)
        put(C_WTUNIT, row, "kg")      # S وحدة الوزن
        print("row %d: P%s %s  %s %sg -> %s SAR  [%s]" % (row, p.get("folder"), p["name"], p["karat"], p["weight"], price, cat))
        row += 1
        ok += 1

    wb.save(OUT)
    print("-" * 64)
    print("Wrote %d product rows -> %s  (VAT %s)" % (ok, OUT, VAT_RATE))
    if failed:
        print("%d failed." % failed)
        sys.exit(2)


if __name__ == "__main__":
    main()
